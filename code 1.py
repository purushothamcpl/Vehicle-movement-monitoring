import cv2
import imutils
import pytesseract
import winsound
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Set the Tesseract OCR executable path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Function to check if a string exists in a file
def check_if_string_in_file(file_name, string_to_search):
    try:
        with open(file_name, 'r') as read_obj:
            for line in read_obj:
                if string_to_search in line:
                    return True
    except FileNotFoundError:
        return False
    return False

# Function to process an image and detect the number plate
def process_image(image_path):
    image = cv2.imread(image_path)
    if image is None:
        print(f"Error: Could not open or read the image file at {image_path}")
        return None, None, None

    # Resize and standardize the image
    image = imutils.resize(image, width=500)
    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Reduce noise and smooth the image
    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    # Find edges in the image
    edged = cv2.Canny(gray, 170, 200)

    # Find contours based on the edges
    cnts, _ = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:30]
    NumberPlateCount = None

    # Loop over contours to find the best possible contour of the number plate
    for i in cnts:
        perimeter = cv2.arcLength(i, True)
        approx = cv2.approxPolyDP(i, 0.02 * perimeter, True)
        if len(approx) == 4:
            NumberPlateCount = approx
            x, y, w, h = cv2.boundingRect(i)
            crp_img = image[y:y + h, x:x + w]
            return crp_img, image, gray

    return None, image, gray

# Directory containing images of vehicles
dataset_dir = 'Dataset'
image_files = [f for f in os.listdir(dataset_dir) if os.path.isfile(os.path.join(dataset_dir, f))]

# Path to the database file
database_file = './Database/Database.txt'

# Path to the Excel file for saving detection logs
excel_path = 'detection_log.xlsx'

# Function to format duration in "days, hours, minutes"
def format_duration(duration):
    days = duration.days
    hours, remainder = divmod(duration.seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{days} days, {hours} hours, {minutes} minutes"

# Function to get the next available slot
def get_next_available_slot(df):
    used_slots = set(df['Slot'].dropna())
    for slot in range(1, 11):  # Limited to 10 slots
        if slot not in used_slots:
            return slot
    return None

# Function to clear the existing data in the Excel file
def clear_excel_file():
    df = pd.DataFrame(columns=['Number Plate', 'Entry Date', 'Entry Time', 'Exit Date', 'Exit Time', 'Duration', 'Status', 'Slot'])
    df.to_excel(excel_path, index=False)

# Function to update or add data in the Excel file
def update_excel(entry_data, exit_data=None):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        max_row = ws.max_row

        # Read the existing DataFrame
        df = pd.read_excel(excel_path)

        plate_exists = False
        for row in range(2, max_row + 1):  # Assuming the first row is the header
            if ws.cell(row=row, column=1).value == entry_data['Number Plate']:
                if ws.cell(row=row, column=4).value and ws.cell(row=row, column=5).value:
                    # Calculate the duration between entry and exit times
                    entry_datetime_str = f"{ws.cell(row=row, column=2).value} {ws.cell(row=row, column=3).value}"
                    exit_datetime_str = f"{ws.cell(row=row, column=4).value} {ws.cell(row=row, column=5).value}"
                    entry_datetime = datetime.strptime(entry_datetime_str, "%Y-%m-%d %H:%M:%S")
                    exit_datetime = datetime.strptime(exit_datetime_str, "%Y-%m-%d %H:%M:%S")
                    duration = exit_datetime - entry_datetime
                    duration_str = format_duration(duration)
                    ws.cell(row=row, column=6, value=duration_str)
                    ws.cell(row=row, column=7, value="Outside")
                    ws.cell(row=row, column=8, value='')  # Clear the slot number
                    continue  # Skip if exit date and time are filled
                if exit_data:
                    ws.cell(row=row, column=4, value=exit_data['Exit Date'])
                    ws.cell(row=row, column=5, value=exit_data['Exit Time'])
                    # Calculate the duration between entry and exit times
                    entry_datetime_str = f"{ws.cell(row=row, column=2).value} {ws.cell(row=row, column=3).value}"
                    exit_datetime_str = f"{exit_data['Exit Date']} {exit_data['Exit Time']}"
                    entry_datetime = datetime.strptime(entry_datetime_str, "%Y-%m-%d %H:%M:%S")
                    exit_datetime = datetime.strptime(exit_datetime_str, "%Y-%m-%d %H:%M:%S")
                    duration = exit_datetime - entry_datetime
                    duration_str = format_duration(duration)
                    ws.cell(row=row, column=6, value=duration_str)
                    ws.cell(row=row, column=7, value="Outside")
                    ws.cell(row=row, column=8, value='')  # Clear the slot number
                plate_exists = True
                break

        if not plate_exists or exit_data is None:
            # Append new row for entry data
            slot_number = get_next_available_slot(df)
            if slot_number is not None:
                ws.append([entry_data['Number Plate'], entry_data['Entry Date'], entry_data['Entry Time'], '', '', '', 'Inside', slot_number])
            else:
                print("All parking slots are occupied. Cannot allow entry.")
                return False  # Indicate that entry is not allowed

        wb.save(excel_path)
    else:
        # Create a new Excel file with headers
        df = pd.DataFrame(columns=['Number Plate', 'Entry Date', 'Entry Time', 'Exit Date', 'Exit Time', 'Duration', 'Status', 'Slot'])
        slot_number = get_next_available_slot(df)
        if slot_number is not None:
            entry_data['Slot'] = slot_number
            new_row = pd.DataFrame([entry_data])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(excel_path, index=False)
        else:
            print("All parking slots are occupied. Cannot allow entry.")
            return False  # Indicate that entry is not allowed

    return True  # Indicate that entry is allowed

# Clear the existing data in the Excel file before running the script
clear_excel_file()

# Dictionary to store the latest status of each number plate
latest_status = {}

# Process each image in the dataset
for image_file in image_files:
    image_path = os.path.join(dataset_dir, image_file)
    cropped_image, original_image, gray_image = process_image(image_path)

    if cropped_image is not None:
        text = pytesseract.image_to_string(cropped_image, lang='eng').strip()
        print(f"Detected Text from {image_file}: {text}")
        text = ''.join(e for e in text if e.isalnum())

        if text:
            is_registered = check_if_string_in_file(database_file, text)

            mod_time = os.path.getmtime(image_path)
            mod_datetime = datetime.fromtimestamp(mod_time)
            current_date = mod_datetime.strftime("%Y-%m-%d")
            current_time = mod_datetime.strftime("%H:%M:%S")

            print(f"Date and Time for {image_file}: {current_date} {current_time}")

            if is_registered:
                print(f'Number Plate {text} is Registered')
                winsound.Beep(2500, 1200)

                entry_data = {
                    'Number Plate': text,
                    'Entry Date': current_date,
                    'Entry Time': current_time
                }

                exit_data = {
                    'Number Plate': text,
                    'Exit Date': current_date,
                    'Exit Time': current_time
                }

                car_entering = True  # Assume the car is entering by default

                if os.path.exists(excel_path):
                    existing_df = pd.read_excel(excel_path)
                    if text in existing_df['Number Plate'].values:
                        # Check if the car is exiting
                        row = existing_df.loc[existing_df['Number Plate'] == text].iloc[-1]
                        if pd.isna(row['Exit Date']) and pd.isna(row['Exit Time']):
                            car_entering = False  # The car is exiting
                            if not update_excel(entry_data, exit_data):
                                cv2.putText(original_image, "No space inside to park", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                                cv2.imshow("Car Status", original_image)
                                cv2.imshow("Grayscale Image", gray_image)
                                cv2.waitKey(0)
                                cv2.destroyAllWindows()
                                continue
                            car_entering = False
                        else:
                            if not update_excel(entry_data):
                                cv2.putText(original_image, "No space inside to park", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                                cv2.imshow("Car Status", original_image)
                                cv2.imshow("Grayscale Image", gray_image)
                                cv2.waitKey(0)
                                cv2.destroyAllWindows()
                                continue
                    else:
                        if not update_excel(entry_data):
                            cv2.putText(original_image, "No space inside to park", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            cv2.imshow("Car Status", original_image)
                            cv2.imshow("Grayscale Image", gray_image)
                            cv2.waitKey(0)
                            cv2.destroyAllWindows()
                            continue
                else:
                    if not update_excel(entry_data):
                        cv2.putText(original_image, "No space inside to park", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                        cv2.imshow("Car Status", original_image)
                        cv2.imshow("Grayscale Image", gray_image)
                        cv2.waitKey(0)
                        cv2.destroyAllWindows()
                        continue

                status = "Outside" if exit_data and not car_entering else "Inside"
                latest_status[text] = (image_file, status)

                # Display the car image with "Allowed" text only if the car is entering
                if car_entering:
                    cv2.putText(original_image, "Allowed", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                    cv2.imshow("Car Status", original_image)
                    cv2.imshow("Grayscale Image", gray_image)
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
            else:
                print(f'Number Plate {text} is Not Registered')
                # Display the car image with "Not Allowed" text
                cv2.putText(original_image, "Not Allowed", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                cv2.imshow("Car Status", original_image)
                cv2.imshow("Grayscale Image", gray_image)
                cv2.waitKey(0)
                cv2.destroyAllWindows()
        else:
            print(f"No valid text found in image {image_file}")
    else:
        print(f"Error: No valid number plate contour found for image {image_file}")

# Read the latest status from the Excel file
if os.path.exists(excel_path):
    df = pd.read_excel(excel_path)
    for index, row in df.iterrows():
        number_plate = row['Number Plate']
        status = row['Status']
        slot = row['Slot']
        image_file = latest_status[number_plate][0] if number_plate in latest_status else None
        if image_file:
            latest_status[number_plate] = (image_file, status, slot)

